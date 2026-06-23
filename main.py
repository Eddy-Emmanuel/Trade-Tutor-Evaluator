import streamlit as st
import pandas as pd
import numpy as np
import io
from openpyxl.utils import get_column_letter

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SMA Trading Signal Calculator",
    page_icon="📈",
    layout="wide",
)

# ── Styling ─────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .metric-card {
        background: #1e2130;
        border: 1px solid #2d3250;
        border-radius: 10px;
        padding: 16px 20px;
        text-align: center;
    }
    .metric-label { color: #8892b0; font-size: 0.78rem; letter-spacing: 0.08em; text-transform: uppercase; margin-bottom: 4px; }
    .metric-value { color: #ccd6f6; font-size: 1.6rem; font-weight: 700; }
    .metric-value.green { color: #64ffda; }
    .metric-value.red   { color: #ff6b6b; }
    .section-header {
        font-size: 1.05rem; font-weight: 600; color: #ccd6f6;
        border-left: 3px solid #64ffda; padding-left: 10px;
        margin: 20px 0 10px;
    }
    div[data-testid="stExpander"] { border: 1px solid #2d3250; border-radius: 8px; }
</style>
""", unsafe_allow_html=True)

# ── Helpers ─────────────────────────────────────────────────────────────────────

def parse_timestamp(series: pd.Series) -> pd.Series:
    """Convert Excel serial numbers or datetime strings to pandas Timestamps."""
    def _convert(v):
        if pd.isna(v):
            return pd.NaT
        try:
            fv = float(v)
            # Excel date serial: days since 1899-12-30
            return pd.Timestamp("1899-12-30") + pd.Timedelta(days=fv)
        except (ValueError, TypeError):
            pass
        return pd.to_datetime(v, errors="coerce")
    return series.apply(_convert)


def load_file(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file, engine="openpyxl")

    # Normalise column names
    df.columns = [c.strip() for c in df.columns]
    rename = {}
    for c in df.columns:
        cl = c.lower().replace(" ", "_")
        if cl in ("s/n", "sn", "s_n", "#"):
            rename[c] = "S/N"
        elif cl == "symbol":
            rename[c] = "Symbol"
        elif "transaction" in cl and "time" in cl:
            rename[c] = "Transaction Time"
        elif "mid" in cl and "price" in cl:
            rename[c] = "Mid Price"
        elif "moving" in cl and "average" in cl:
            rename[c] = "Moving Average"
        elif cl == "position":
            rename[c] = "Position"
        elif cl == "action":
            rename[c] = "Action"
    df = df.rename(columns=rename)

    # Keep only expected columns (drop extra annotation columns Excel may have)
    keep = [c for c in ["S/N", "Symbol", "Transaction Time", "Mid Price",
                         "Moving Average", "Position", "Action"] if c in df.columns]
    df = df[keep].copy()

    # Parse & clean
    if "Transaction Time" in df.columns:
        df["Transaction Time"] = parse_timestamp(df["Transaction Time"])
    if "Mid Price" in df.columns:
        df["Mid Price"] = pd.to_numeric(df["Mid Price"], errors="coerce")

    df = df.dropna(subset=["Mid Price"]).reset_index(drop=True)
    return df


def compute_sma(prices: pd.Series, window: int) -> pd.Series:
    """Rolling SMA – returns NaN for the first (window-1) rows."""
    return prices.rolling(window=window).mean()


def compute_thresholds(sma: pd.Series, buy_pct: float, sell_pct: float,
                        buy_direction: str, sell_direction: str):
    buy_sign  = +1 if buy_direction  == "above" else -1
    sell_sign = +1 if sell_direction == "above" else -1
    buy_thresh  = sma * (1 + buy_sign  * buy_pct  / 100)
    sell_thresh = sma * (1 + sell_sign * sell_pct / 100)
    return buy_thresh, sell_thresh


def compute_signals(prices: pd.Series, buy_thresh: pd.Series, sell_thresh: pd.Series,
                     buy_direction: str, sell_direction: str):
    """State-machine: Out → In on Buy, In → Out on Sell."""
    position = []
    action   = []
    buy_cond = []
    sell_cond = []
    state = "Out"

    for i in range(len(prices)):
        p  = prices.iloc[i]
        bt = buy_thresh.iloc[i]
        st = sell_thresh.iloc[i]

        if pd.isna(bt) or pd.isna(st):
            bc = False
            sc = False
        else:
            bc = (p > bt) if buy_direction  == "above" else (p < bt)
            sc = (p < st) if sell_direction == "below" else (p > st)

        buy_cond.append(bc)
        sell_cond.append(sc)

        if state == "Out" and bc:
            state = "In"
            action.append("Buy")
        elif state == "In" and sc:
            state = "Out"
            action.append("Sell")
        else:
            action.append("Hold")
        position.append(state)

    return position, action, buy_cond, sell_cond


def compute_pnl(df_result: pd.DataFrame, shares: int = 500) -> dict:
    """Simple PnL: sum up completed Buy→Sell round-trips."""
    buys  = df_result[df_result["Action"] == "Buy" ]["Mid Price"].tolist()
    sells = df_result[df_result["Action"] == "Sell"]["Mid Price"].tolist()
    pairs = min(len(buys), len(sells))
    pnl_list = [(sells[i] - buys[i]) * shares for i in range(pairs)]
    total = sum(pnl_list)
    wins  = sum(1 for x in pnl_list if x > 0)
    losses= pairs - wins
    return {"total": total, "trades": pairs, "wins": wins, "losses": losses,
            "pnl_list": pnl_list, "buys": buys[:pairs], "sells": sells[:pairs]}


def write_formula_workbook(df_result: pd.DataFrame, window: int,
                            buy_pct: float, sell_pct: float,
                            buy_direction: str, sell_direction: str) -> io.BytesIO:
    """Write df_result to an .xlsx file where the calculated columns contain
    live Excel formulas instead of static pre-computed values."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_result.to_excel(writer, index=False, sheet_name="SMA Results")
        ws = writer.sheets["SMA Results"]

        col_letter = {name: get_column_letter(df_result.columns.get_loc(name) + 1)
                      for name in df_result.columns}

        mid_L = col_letter["Mid Price"]
        ma_L  = col_letter["Moving Average (calc)"]
        bt_L  = col_letter["Buy Threshold"]
        st_L  = col_letter["Sell Threshold"]
        bc_L  = col_letter["Buy Condition"]
        sc_L  = col_letter["Sell Condition"]
        pos_L = col_letter["Position (calc)"]
        act_L = col_letter["Action (calc)"]

        buy_sign  = 1 if buy_direction  == "above" else -1
        sell_sign = 1 if sell_direction == "above" else -1
        buy_op    = ">" if buy_direction  == "above" else "<"
        sell_op   = "<" if sell_direction == "below" else ">"

        n_rows = len(df_result)
        for i in range(n_rows):
            r = i + 2  # account for header row + 1-indexing

            # Moving Average — rolling AVERAGE() over the window
            if i >= window - 1:
                ma_formula = f"=AVERAGE({mid_L}{r - window + 1}:{mid_L}{r})"
            else:
                ma_formula = '=""'
            ws[f"{ma_L}{r}"] = ma_formula

            # Buy / Sell thresholds — % offset from the SMA
            ws[f"{bt_L}{r}"] = (
                f'=IF({ma_L}{r}="","",{ma_L}{r}*(1+({buy_sign})*{buy_pct}/100))'
            )
            ws[f"{st_L}{r}"] = (
                f'=IF({ma_L}{r}="","",{ma_L}{r}*(1+({sell_sign})*{sell_pct}/100))'
            )

            # Buy / Sell conditions — price vs. threshold comparison
            ws[f"{bc_L}{r}"] = (
                f'=IF(OR({ma_L}{r}="",{bt_L}{r}=""),FALSE,{mid_L}{r}{buy_op}{bt_L}{r})'
            )
            ws[f"{sc_L}{r}"] = (
                f'=IF(OR({ma_L}{r}="",{st_L}{r}=""),FALSE,{mid_L}{r}{sell_op}{st_L}{r})'
            )

            # Position / Action — state machine referencing the previous row
            prev_pos = '"Out"' if i == 0 else f"{pos_L}{r - 1}"
            ws[f"{pos_L}{r}"] = (
                f'=IF(AND({prev_pos}="Out",{bc_L}{r}=TRUE),"In",'
                f'IF(AND({prev_pos}="In",{sc_L}{r}=TRUE),"Out",{prev_pos}))'
            )
            ws[f"{act_L}{r}"] = (
                f'=IF(AND({prev_pos}="Out",{bc_L}{r}=TRUE),"Buy",'
                f'IF(AND({prev_pos}="In",{sc_L}{r}=TRUE),"Sell","Hold"))'
            )

    buf.seek(0)
    return buf


# ── App layout ──────────────────────────────────────────────────────────────────
st.title("📈 SMA Trading Signal Calculator")
st.caption("Upload a trade file · configure SMA parameters · inspect buy/sell signals")

# ── Sidebar ──────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Configuration")

    uploaded = st.file_uploader("Upload Excel or CSV", type=["xlsx", "xls", "csv"])

    st.markdown("---")
    st.subheader("SMA Parameters")
    window = st.number_input("Window (periods)", min_value=2, max_value=500, value=2, step=1)

    st.subheader("Buy Rule")
    buy_direction = st.selectbox("Buy when price is", ["above", "below"],
                                  index=0, key="buy_dir")
    buy_pct = st.number_input("Buy threshold %", min_value=0.0, max_value=100.0,
                               value=0.05, step=0.01, format="%.4f", key="buy_pct")

    st.subheader("Sell Rule")
    sell_direction = st.selectbox("Sell when price is", ["below", "above"],
                                   index=0, key="sell_dir")
    sell_pct = st.number_input("Sell threshold %", min_value=0.0, max_value=100.0,
                                value=0.05, step=0.01, format="%.4f", key="sell_pct")

    st.subheader("Trade Size")
    shares = st.number_input("Shares per trade", min_value=1, value=500, step=1)

    st.markdown("---")
    run_btn = st.button("▶ Calculate", use_container_width=True, type="primary")

# ── Main ─────────────────────────────────────────────────────────────────────────
if uploaded is None:
    st.info("👈  Upload an Excel or CSV file to get started.")
    with st.expander("Expected column format"):
        st.markdown("""
| Column | Description |
|---|---|
| `S/N` | Row number |
| `Symbol` | Ticker / instrument |
| `Transaction Time` | Timestamp or Excel serial |
| `Mid Price` | Mid-market price used for calculation |
| `Moving Average` | Pre-computed value *(optional – will be recalculated)* |
| `Position` | `In` / `Out` *(optional)* |
| `Action` | `Buy` / `Sell` / `Hold` *(optional)* |
""")
    st.stop()

# ── Load ─────────────────────────────────────────────────────────────────────────
try:
    df_raw = load_file(uploaded)
except Exception as e:
    st.error(f"Could not parse file: {e}")
    st.stop()

if df_raw.empty or "Mid Price" not in df_raw.columns:
    st.error("File loaded but no valid `Mid Price` column found.")
    st.stop()

st.success(f"Loaded **{len(df_raw):,}** rows · {df_raw['Symbol'].iloc[0] if 'Symbol' in df_raw.columns else ''}")

# ── Calculate ─────────────────────────────────────────────────────────────────────
if run_btn or True:   # auto-run on load; re-runs when button pressed
    prices = df_raw["Mid Price"]
    sma    = compute_sma(prices, window)
    buy_thresh, sell_thresh = compute_thresholds(sma, buy_pct, sell_pct,
                                                  buy_direction, sell_direction)
    position, action, buy_cond, sell_cond = compute_signals(
        prices, buy_thresh, sell_thresh, buy_direction, sell_direction)

    df_result = df_raw.copy()
    df_result["Moving Average (calc)"] = sma.round(7)
    df_result["Buy Threshold"]         = buy_thresh.round(7)
    df_result["Sell Threshold"]        = sell_thresh.round(7)
    df_result["Buy Condition"]         = buy_cond
    df_result["Sell Condition"]        = sell_cond
    df_result["Position (calc)"]       = position
    df_result["Action (calc)"]         = action

    pnl = compute_pnl(df_result, shares)

    # ── KPI row ──────────────────────────────────────────────────────────────────
    col1, col2, col3, col4, col5 = st.columns(5)
    pnl_color = "green" if pnl["total"] >= 0 else "red"
    kpis = [
        (col1, "Total Rows",    f"{len(df_result):,}",          ""),
        (col2, "SMA Window",    f"{window} bars",                ""),
        (col3, "Round Trips",   f"{pnl['trades']}",             ""),
        (col4, "Win / Loss",    f"{pnl['wins']} / {pnl['losses']}", ""),
        (col5, "Est. P&L",      f"£{pnl['total']:,.2f}",         pnl_color),
    ]
    for col, lbl, val, color in kpis:
        with col:
            st.markdown(
                f'<div class="metric-card">'
                f'<div class="metric-label">{lbl}</div>'
                f'<div class="metric-value {color}">{val}</div>'
                f'</div>', unsafe_allow_html=True)

    # ── Data table ───────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">Calculated Results</div>',
                unsafe_allow_html=True)

    display_cols = ["Transaction Time", "Mid Price",
                    "Moving Average (calc)", "Buy Threshold", "Sell Threshold",
                    "Buy Condition", "Sell Condition",
                    "Position (calc)", "Action (calc)"]
    display_cols = [c for c in display_cols if c in df_result.columns]

    def highlight_action(row):
        a = row.get("Action (calc)", "")
        if a == "Buy":
            return ["background-color: #1a3a2a"] * len(row)
        elif a == "Sell":
            return ["background-color: #3a1a1a"] * len(row)
        return [""] * len(row)

    show_df = df_result[display_cols].copy()
    for c in ["Mid Price", "Moving Average (calc)", "Buy Threshold", "Sell Threshold"]:
        if c in show_df.columns:
            show_df[c] = show_df[c].map(lambda x: f"{x:.7f}" if pd.notna(x) else "")

    MAX_STYLE_CELLS = 262_144
    total_cells = show_df.shape[0] * show_df.shape[1]

    if total_cells <= MAX_STYLE_CELLS:
        st.dataframe(show_df.style.apply(highlight_action, axis=1),
                    width='stretch', height=400)
    else:
        st.caption(f"⚠️ Large dataset ({total_cells:,} cells) — row highlighting disabled for performance.")
        st.dataframe(show_df, width='stretch', height=400)

    # ── Trade log ────────────────────────────────────────────────────────────────
    if pnl["trades"] > 0:
        st.markdown('<div class="section-header">Round-Trip Trade Log</div>',
                    unsafe_allow_html=True)
        trades_df = pd.DataFrame({
            "#":         range(1, pnl["trades"] + 1),
            "Buy Price": [f"{b:.7f}" for b in pnl["buys"]],
            "Sell Price":[f"{s:.7f}" for s in pnl["sells"]],
            "P&L (£)":   [f"{p:+.4f}" for p in pnl["pnl_list"]],
        })
        st.dataframe(trades_df, use_container_width=True, hide_index=True)

    # ── Download ─────────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">Export</div>', unsafe_allow_html=True)
    buf = write_formula_workbook(df_result, window, buy_pct, sell_pct,
                                  buy_direction, sell_direction)
    st.download_button("⬇ Download Results (Excel)", data=buf,
                        file_name="sma_results.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")