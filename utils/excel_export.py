"""
utils/excel_export.py
Builds an .xlsx workbook where every calculated cell contains a LIVE Excel
formula (not a hardcoded value). Change Window / % thresholds / direction /
indicator settings on the Settings sheet and the Results sheet recalculates.

Public entry point: build_workbook(...) -> openpyxl.Workbook

QA round 2 — what changed
-------------------------
1. NO HARDCODED CONSTANTS IN FORMULAS. RSI 30/70, Stochastic 20/80, ADX 25/20,
   MACD 12/26/9 (as 2/13, 2/27, 2/10), Fibonacci's 38.2/61.8 anchors and the
   σ denominator were all baked into the formula strings. Every one of them is
   now a labelled Settings cell, so the workbook agrees with whatever the
   sidebar sent AND stays editable in Excel.

2. WARM-UP PARITY. The %K formula used to emit 50 during warm-up, exactly like
   the Python `.fillna(50)`, which produced a phantom crossover on the first
   real row. It now emits "". Every indicator's Buy/Sell Condition is gated by
   the same warm-up length the engine uses, so Position/Action hold until the
   lookback is genuinely satisfied.

3. DIRECTION PARITY. Bollinger, Std Dev, Fibonacci and Stochastic/RSI/ADX
   ignored Buy/Sell Direction entirely — the workbook always ran the canonical
   sense while Python honoured the dropdown. They now read the direction cells,
   and the band indicators gained the Buy/Sell Threshold columns Python emits.

4. ADX PERIOD. ADX read the generic Window cell. It now reads its own ADX
   Period cell, which is what the sidebar sends.

Warm-up lengths mirror indicators/engine.py exactly (row r holds 0-based index
i = r-2, and a signal is blocked while i < warmup):
    SMA / EMA / Bollinger / StdDev / Fib : window - 1
    Stochastic                           : window
    RSI                                  : window + 1
    MACD                                 : slow + signal - 1
    ADX                                  : 2 * adx_period
    Heikin Ashi                          : 1
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BLUE = Font(color="0000FF")
BLACK = Font(color="000000")
GREY = Font(color="808080", italic=True)
HDR_FILL = PatternFill("solid", start_color="1F2937", end_color="1F2937")
HDR_FONT = Font(color="FFFFFF", bold=True)

RAW_COL_ORDER = [
    "S/N", "Symbol", "Transaction Time",
    "Ask Price", "Bid Price", "Mid Price",
    "Moving Average", "Position", "Action",
]

# ── Fixed Settings cells (rows 1-8). Indicator settings start at row 9. ─────
WIN = "Settings!$B$1"
BUYPCT = "Settings!$B$2"
SELLPCT = "Settings!$B$3"
BUYDIR = "Settings!$B$4"
SELLDIR = "Settings!$B$5"
ALPHA = "Settings!$B$6"
FIRST_PARAM_ROW = 9

# Labels for the dynamic rows, keyed by the engine's param name.
PARAM_LABELS = {
    "oversold": "Buy Threshold",
    "overbought": "Sell Threshold",
    "strong": "Buy Threshold (ADX)",
    "weak": "Sell Threshold (ADX)",
    "d_window": "%D Period",
    "adx_window": "ADX Period",
    "fast_window": "Fast EMA Period",
    "slow_window": "Slow EMA Period",
    "signal_window": "Signal EMA Period",
    "k": "Band k",
    "ddof": "Sigma denominator (0=population, 1=sample)",
    "buy_level": "Buy anchor level",
    "sell_level": "Sell anchor level",
    "transition_only": "Signal on colour change only",
}


def _build_settings_sheet(
    wb, window, buy_pct, sell_pct, buy_direction, sell_direction,
    repeat_flag=False, params=None, indicator_name="",
):
    """Write the fixed rows, then one labelled row per indicator setting.

    Returns S: {param_key -> absolute cell ref}, so the formula builders can
    point at a cell instead of embedding a literal.
    """
    params = params or {}
    ws = wb.active
    ws.title = "Settings"

    # These three read their own named periods, so the generic Window is inert.
    win_label = "Window (Periods)"
    if indicator_name in ("MACD", "ADX", "Heikin Ashi"):
        win_label = f"Window (Periods) — not used by {indicator_name}"

    fixed = [
        (win_label, window),
        ("Buy %", buy_pct),
        ("Sell %", sell_pct),
        ("Buy Direction", buy_direction),
        ("Sell Direction", sell_direction),
        ("EMA alpha (auto)", "=2/(B1+1)"),
        ("Repeat Trade Flag", "TRUE" if repeat_flag else "FALSE"),
        ("Indicator", indicator_name),
    ]
    for i, (label, val) in enumerate(fixed, start=1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=val)
        c.font = BLACK if (isinstance(val, str) and val.startswith("=")) else BLUE

    S = {}
    row = FIRST_PARAM_ROW
    for key, value in params.items():
        label = PARAM_LABELS.get(key, key)
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        if isinstance(value, bool):
            value = "TRUE" if value else "FALSE"
        ws.cell(row=row, column=2, value=value).font = BLUE
        S[key] = f"Settings!$B${row}"
        row += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws["D1"] = (
        "Edit the blue cells — the Results sheet recalculates automatically. "
        "Direction must be exactly 'above' or 'below'. Repeat Trade Flag is "
        "baked into the Position/Action formulas at export time; changing it "
        "here will NOT update them, so re-export instead."
    )
    ws["D1"].font = GREY
    return S


def _offset(price_cell, win_ref=WIN):
    """Trailing rolling window range ending at price_cell, length = win_ref."""
    return f"OFFSET({price_cell},-({win_ref}-1),0,{win_ref},1)"


def _gate(warmup_expr: str, formula_body: str) -> str:
    """Return FALSE while the lookback is unsatisfied.

    Row r holds 0-based index i = r-2, matching engine._blank_warmup.
    """
    return f'=IF((ROW()-2)<({warmup_expr}),FALSE,{formula_body})'


def _cross(prev, curr, level_ref, dir_ref):
    """Direction-literal crossing test — mirrors engine._edge_cross."""
    return (f'IF({dir_ref}="above",'
            f'AND({prev}<={level_ref},{curr}>{level_ref}),'
            f'AND({prev}>={level_ref},{curr}<{level_ref}))')


def _cross_guarded(prev, curr, level_ref, dir_ref, warmup_expr):
    """Crossing test that refuses to fire on a blank predecessor or value."""
    body = f'IF(OR({prev}="",{curr}=""),FALSE,{_cross(prev, curr, level_ref, dir_ref)})'
    return _gate(warmup_expr, body)


def _state_machine_formulas(r, buy_cell, sell_cell, prev_pos_ref, repeat_flag=False):
    """Universal Position(calc)/Action(calc) formulas (used by ALL indicators).

    Mirrors indicators.engine._state_machine.
    """
    if repeat_flag:
        action = f'=IF({buy_cell},"Buy",IF({sell_cell},"Sell","Hold"))'
        position = f'=IF({buy_cell},"In",IF({sell_cell},"Out",{prev_pos_ref}))'
    else:
        action = (f'=IF(AND({prev_pos_ref}="Out",{buy_cell}),"Buy",'
                  f'IF(AND({prev_pos_ref}="In",{sell_cell}),"Sell","Hold"))')
        position = (f'=IF(AND({prev_pos_ref}="Out",{buy_cell}),"In",'
                    f'IF(AND({prev_pos_ref}="In",{sell_cell}),"Out",{prev_pos_ref}))')
    return position, action


def _status_formula(uploaded_action_cell, action_calc_cell):
    return (f'=IF(TRIM({uploaded_action_cell})="","N/A",'
            f'IF(LOWER(TRIM({uploaded_action_cell}))=LOWER(TRIM({action_calc_cell})),'
            f'"Pass","Fail"))')


# ─────────────────────────── per-indicator column specs ─────────────────────
# Each factory takes S (settings refs) and returns a list of
# (col_name, hidden, formula_fn(r, L) -> str_or_value).
# L is a dict {col_name: column_letter}; "price" always maps to the price col.

def _sma_spec(S):
    warm = f"{WIN}-1"

    def ma(r, L):
        p = f'{L["price"]}{r}'
        return f'=IF((ROW()-1)<{WIN},"",AVERAGE({_offset(p)}))'

    def buyt(r, L):
        m = f'{L["Moving Average (calc)"]}{r}'
        return f'=IF({m}="","",{m}*(1+IF({BUYDIR}="above",1,-1)*{BUYPCT}/100))'

    def sellt(r, L):
        m = f'{L["Moving Average (calc)"]}{r}'
        return f'=IF({m}="","",{m}*(1+IF({SELLDIR}="above",1,-1)*{SELLPCT}/100))'

    def buyc(r, L):
        p = f'{L["price"]}{r}'; bt = f'{L["Buy Threshold"]}{r}'
        return _gate(warm, f'IF({bt}="",FALSE,IF({BUYDIR}="above",{p}>{bt},{p}<{bt}))')

    def sellc(r, L):
        p = f'{L["price"]}{r}'; st = f'{L["Sell Threshold"]}{r}'
        return _gate(warm, f'IF({st}="",FALSE,IF({SELLDIR}="below",{p}<{st},{p}>{st}))')

    return [
        ("Moving Average (calc)", False, ma),
        ("Buy Threshold", False, buyt),
        ("Sell Threshold", False, sellt),
        ("Buy Condition", False, buyc),
        ("Sell Condition", False, sellc),
    ]


def _ema_spec(S):
    warm = f"{WIN}-1"

    def ema(r, L):
        if r == 2:
            return f'={L["price"]}{r}'
        p = f'{L["price"]}{r}'; prev = f'{L["EMA (calc)"]}{r-1}'
        return f'={ALPHA}*{p}+(1-{ALPHA})*{prev}'

    def buyt(r, L):
        m = f'{L["EMA (calc)"]}{r}'
        return f'={m}*(1+IF({BUYDIR}="above",1,-1)*{BUYPCT}/100)'

    def sellt(r, L):
        m = f'{L["EMA (calc)"]}{r}'
        return f'={m}*(1+IF({SELLDIR}="above",1,-1)*{SELLPCT}/100)'

    def buyc(r, L):
        p = f'{L["price"]}{r}'; bt = f'{L["Buy Threshold"]}{r}'
        return _gate(warm, f'IF({BUYDIR}="above",{p}>{bt},{p}<{bt})')

    def sellc(r, L):
        p = f'{L["price"]}{r}'; st = f'{L["Sell Threshold"]}{r}'
        return _gate(warm, f'IF({SELLDIR}="below",{p}<{st},{p}>{st})')

    return [
        ("EMA (calc)", False, ema),
        ("Buy Threshold", False, buyt),
        ("Sell Threshold", False, sellt),
        ("Buy Condition", False, buyc),
        ("Sell Condition", False, sellc),
    ]


def _stochastic_spec(S):
    warm = WIN                      # engine: first_valid(%K) + 1 = window
    dwin = S.get("d_window", "3")
    over_s = S.get("oversold", "20")
    over_b = S.get("overbought", "80")

    def k(r, L):
        p = f'{L["price"]}{r}'
        off = _offset(p)
        lo, hi = f'MIN({off})', f'MAX({off})'
        # No placeholder 50 — an undefined %K stays blank so it cannot cross.
        return (f'=IF((ROW()-1)<{WIN},"",'
                f'IF(({hi}-{lo})=0,"",({p}-{lo})/({hi}-{lo})*100))')

    def d(r, L):
        kc = f'{L["%K (calc)"]}{r}'
        return (f'=IF((ROW()-1)<({WIN}+{dwin}-1),"",'
                f'AVERAGE(OFFSET({kc},-({dwin}-1),0,{dwin},1)))')

    def buyc(r, L):
        if r == 2:
            return "=FALSE"
        return _cross_guarded(f'{L["%K (calc)"]}{r-1}', f'{L["%K (calc)"]}{r}',
                              over_s, BUYDIR, warm)

    def sellc(r, L):
        if r == 2:
            return "=FALSE"
        return _cross_guarded(f'{L["%K (calc)"]}{r-1}', f'{L["%K (calc)"]}{r}',
                              over_b, SELLDIR, warm)

    return [
        ("%K (calc)", False, k),
        ("%D (Signal)", False, d),
        ("Buy Condition", False, buyc),
        ("Sell Condition", False, sellc),
    ]


def _macd_spec(S):
    fast = S.get("fast_window", "12")
    slow = S.get("slow_window", "26")
    sig = S.get("signal_window", "9")
    warm = f"{slow}+{sig}-1"

    def ema_fast(r, L):
        if r == 2:
            return f'={L["price"]}{r}'
        p = f'{L["price"]}{r}'; prev = f'{L["EMA Fast (helper)"]}{r-1}'
        return f'=(2/({fast}+1))*{p}+(1-(2/({fast}+1)))*{prev}'

    def ema_slow(r, L):
        if r == 2:
            return f'={L["price"]}{r}'
        p = f'{L["price"]}{r}'; prev = f'{L["EMA Slow (helper)"]}{r-1}'
        return f'=(2/({slow}+1))*{p}+(1-(2/({slow}+1)))*{prev}'

    def macd(r, L):
        return f'={L["EMA Fast (helper)"]}{r}-{L["EMA Slow (helper)"]}{r}'

    def signal(r, L):
        m = f'{L["MACD (calc)"]}{r}'
        if r == 2:
            return f'={m}'
        prev = f'{L["MACD Signal"]}{r-1}'
        return f'=(2/({sig}+1))*{m}+(1-(2/({sig}+1)))*{prev}'

    def buyt(r, L):
        s = f'{L["MACD Signal"]}{r}'
        return f'={s}*(1+IF({BUYDIR}="above",1,-1)*{BUYPCT}/100)'

    def sellt(r, L):
        s = f'{L["MACD Signal"]}{r}'
        return f'={s}*(1+IF({SELLDIR}="above",1,-1)*{SELLPCT}/100)'

    def hist(r, L):
        return f'={L["MACD (calc)"]}{r}-{L["MACD Signal"]}{r}'

    def buyc(r, L):
        if r == 2:
            return "=FALSE"
        pm, cm = f'{L["MACD (calc)"]}{r-1}', f'{L["MACD (calc)"]}{r}'
        pt, ct = f'{L["Buy Threshold"]}{r-1}', f'{L["Buy Threshold"]}{r}'
        return _gate(warm, f'AND({pm}<={pt},{cm}>{ct})')

    def sellc(r, L):
        if r == 2:
            return "=FALSE"
        pm, cm = f'{L["MACD (calc)"]}{r-1}', f'{L["MACD (calc)"]}{r}'
        pt, ct = f'{L["Sell Threshold"]}{r-1}', f'{L["Sell Threshold"]}{r}'
        return _gate(warm, f'AND({pm}>={pt},{cm}<{ct})')

    return [
        ("EMA Fast (helper)", True, ema_fast),
        ("EMA Slow (helper)", True, ema_slow),
        ("MACD (calc)", False, macd),
        ("MACD Signal", False, signal),
        ("MACD Histogram", False, hist),
        ("Buy Threshold", False, buyt),
        ("Sell Threshold", False, sellt),
        ("Buy Condition", False, buyc),
        ("Sell Condition", False, sellc),
    ]


def _std_formula(rng, ddof_ref):
    """STDEVP or STDEV depending on the σ denominator setting."""
    return f'IF({ddof_ref}=0,STDEVP({rng}),STDEV({rng}))'


def _bollinger_spec(S):
    warm = f"{WIN}-1"
    kband = S.get("k", "2")
    ddof = S.get("ddof", "0")

    def mid(r, L):
        p = f'{L["price"]}{r}'
        return f'=IF((ROW()-1)<{WIN},"",AVERAGE({_offset(p)}))'

    def upper(r, L):
        p = f'{L["price"]}{r}'; m = f'{L["BB Middle (calc)"]}{r}'
        return f'=IF({m}="","",{m}+{kband}*{_std_formula(_offset(p), ddof)})'

    def lower(r, L):
        p = f'{L["price"]}{r}'; m = f'{L["BB Middle (calc)"]}{r}'
        return f'=IF({m}="","",{m}-{kband}*{_std_formula(_offset(p), ddof)})'

    def buyt(r, L):
        lo = f'{L["BB Lower"]}{r}'
        return f'=IF({lo}="","",{lo}*(1+IF({BUYDIR}="above",1,-1)*{BUYPCT}/100))'

    def sellt(r, L):
        up = f'{L["BB Upper"]}{r}'
        return f'=IF({up}="","",{up}*(1+IF({SELLDIR}="above",1,-1)*{SELLPCT}/100))'

    def buyc(r, L):
        p = f'{L["price"]}{r}'; bt = f'{L["Buy Threshold"]}{r}'
        return _gate(warm, f'IF({bt}="",FALSE,IF({BUYDIR}="below",{p}<{bt},{p}>{bt}))')

    def sellc(r, L):
        p = f'{L["price"]}{r}'; st = f'{L["Sell Threshold"]}{r}'
        return _gate(warm, f'IF({st}="",FALSE,IF({SELLDIR}="above",{p}>{st},{p}<{st}))')

    return [
        ("BB Middle (calc)", False, mid),
        ("BB Upper", False, upper),
        ("BB Lower", False, lower),
        ("Buy Threshold", False, buyt),
        ("Sell Threshold", False, sellt),
        ("Buy Condition", False, buyc),
        ("Sell Condition", False, sellc),
    ]


def _rsi_spec(S):
    warm = f"{WIN}+1"               # engine: first_valid(RSI) + 1
    over_s = S.get("oversold", "30")
    over_b = S.get("overbought", "70")

    def gain(r, L):
        if r == 2:
            return '=""'
        return f'=MAX({L["price"]}{r}-{L["price"]}{r-1},0)'

    def loss(r, L):
        if r == 2:
            return '=""'
        return f'=MAX({L["price"]}{r-1}-{L["price"]}{r},0)'

    def avgg(r, L):
        off = _offset(f'{L["Gain (helper)"]}{r}')
        return f'=IF(COUNT({off})<{WIN},"",AVERAGE({off}))'

    def avgl(r, L):
        off = _offset(f'{L["Loss (helper)"]}{r}')
        return f'=IF(COUNT({off})<{WIN},"",AVERAGE({off}))'

    def rsi(r, L):
        ag = f'{L["Avg Gain (helper)"]}{r}'; al = f'{L["Avg Loss (helper)"]}{r}'
        return (f'=IF(OR({ag}="",{al}=""),"",'
                f'IF({al}=0,"",100-100/(1+{ag}/{al})))')

    def buyc(r, L):
        if r == 2:
            return "=FALSE"
        return _cross_guarded(f'{L["RSI (calc)"]}{r-1}', f'{L["RSI (calc)"]}{r}',
                              over_s, BUYDIR, warm)

    def sellc(r, L):
        if r == 2:
            return "=FALSE"
        return _cross_guarded(f'{L["RSI (calc)"]}{r-1}', f'{L["RSI (calc)"]}{r}',
                              over_b, SELLDIR, warm)

    return [
        ("Gain (helper)", True, gain),
        ("Loss (helper)", True, loss),
        ("Avg Gain (helper)", True, avgg),
        ("Avg Loss (helper)", True, avgl),
        ("RSI (calc)", False, rsi),
        ("Buy Condition", False, buyc),
        ("Sell Condition", False, sellc),
    ]


def _fibonacci_spec(S):
    warm = f"{WIN}-1"
    buy_lv = S.get("buy_level", "0.382")
    sell_lv = S.get("sell_level", "0.618")

    def lo(r, L):
        p = f'{L["price"]}{r}'
        return f'=IF((ROW()-1)<{WIN},"",MIN({_offset(p)}))'

    def hi(r, L):
        p = f'{L["price"]}{r}'
        return f'=IF((ROW()-1)<{WIN},"",MAX({_offset(p)}))'

    def fib(mult):
        def f(r, L):
            l = f'{L["Roll Low (helper)"]}{r}'; h = f'{L["Roll High (helper)"]}{r}'
            return f'=IF(OR({l}="",{h}=""),"",{l}+{mult}*({h}-{l}))'
        return f

    def buyt(r, L):
        l = f'{L["Roll Low (helper)"]}{r}'; h = f'{L["Roll High (helper)"]}{r}'
        anchor = f'({l}+{buy_lv}*({h}-{l}))'
        return (f'=IF(OR({l}="",{h}=""),"",'
                f'{anchor}*(1+IF({BUYDIR}="above",1,-1)*{BUYPCT}/100))')

    def sellt(r, L):
        l = f'{L["Roll Low (helper)"]}{r}'; h = f'{L["Roll High (helper)"]}{r}'
        anchor = f'({l}+{sell_lv}*({h}-{l}))'
        return (f'=IF(OR({l}="",{h}=""),"",'
                f'{anchor}*(1+IF({SELLDIR}="above",1,-1)*{SELLPCT}/100))')

    def buyc(r, L):
        p = f'{L["price"]}{r}'; bt = f'{L["Buy Threshold"]}{r}'
        return _gate(warm, f'IF({bt}="",FALSE,IF({BUYDIR}="below",{p}<{bt},{p}>{bt}))')

    def sellc(r, L):
        p = f'{L["price"]}{r}'; st = f'{L["Sell Threshold"]}{r}'
        return _gate(warm, f'IF({st}="",FALSE,IF({SELLDIR}="above",{p}>{st},{p}<{st}))')

    return [
        ("Roll Low (helper)", True, lo),
        ("Roll High (helper)", True, hi),
        ("Fib 23.6%", False, fib(0.236)),
        ("Fib 38.2%", False, fib(0.382)),
        ("Fib 50% (calc)", False, fib(0.500)),
        ("Fib 61.8%", False, fib(0.618)),
        ("Fib 78.6%", False, fib(0.786)),
        ("Buy Threshold", False, buyt),
        ("Sell Threshold", False, sellt),
        ("Buy Condition", False, buyc),
        ("Sell Condition", False, sellc),
    ]


def _stddev_spec(S):
    warm = f"{WIN}-1"
    kband = S.get("k", "2")
    ddof = S.get("ddof", "0")
    SIG = "Std Dev σ (calc)"        # single space — matches the engine's column

    def sigma(r, L):
        p = f'{L["price"]}{r}'
        return f'=IF((ROW()-1)<{WIN},"",{_std_formula(_offset(p), ddof)})'

    def mu(r, L):
        p = f'{L["price"]}{r}'
        return f'=IF((ROW()-1)<{WIN},"",AVERAGE({_offset(p)}))'

    def lower(r, L):
        m = f'{L["StdDev Mean"]}{r}'; s = f'{L[SIG]}{r}'
        return f'=IF(OR({m}="",{s}=""),"",{m}-{kband}*{s})'

    def upper(r, L):
        m = f'{L["StdDev Mean"]}{r}'; s = f'{L[SIG]}{r}'
        return f'=IF(OR({m}="",{s}=""),"",{m}+{kband}*{s})'

    def buyt(r, L):
        lo = f'{L["StdDev Lower"]}{r}'
        return f'=IF({lo}="","",{lo}*(1+IF({BUYDIR}="above",1,-1)*{BUYPCT}/100))'

    def sellt(r, L):
        up = f'{L["StdDev Upper"]}{r}'
        return f'=IF({up}="","",{up}*(1+IF({SELLDIR}="above",1,-1)*{SELLPCT}/100))'

    def buyc(r, L):
        p = f'{L["price"]}{r}'; bt = f'{L["Buy Threshold"]}{r}'
        return _gate(warm, f'IF({bt}="",FALSE,IF({BUYDIR}="below",{p}<{bt},{p}>{bt}))')

    def sellc(r, L):
        p = f'{L["price"]}{r}'; st = f'{L["Sell Threshold"]}{r}'
        return _gate(warm, f'IF({st}="",FALSE,IF({SELLDIR}="above",{p}>{st},{p}<{st}))')

    return [
        (SIG, False, sigma),
        ("StdDev Mean", False, mu),
        ("StdDev Lower", False, lower),
        ("StdDev Upper", False, upper),
        ("Buy Threshold", False, buyt),
        ("Sell Threshold", False, sellt),
        ("Buy Condition", False, buyc),
        ("Sell Condition", False, sellc),
    ]


def _adx_spec(S):
    # ADX reads its own period cell, not the generic Window.
    n = S.get("adx_window", "14")
    warm = f"2*{n}"
    strong = S.get("strong", "25")
    weak = S.get("weak", "20")

    def delta(r, L):
        if r == 2:
            return '=""'
        return f'={L["price"]}{r}-{L["price"]}{r-1}'

    def plusdm(r, L):
        d = f'{L["Delta (helper)"]}{r}'
        return f'=IF({d}="","",MAX({d},0))'

    def minusdm(r, L):
        d = f'{L["Delta (helper)"]}{r}'
        return f'=IF({d}="","",MAX(-{d},0))'

    def tr(r, L):
        d = f'{L["Delta (helper)"]}{r}'
        return f'=IF({d}="","",ABS({d}))'

    def _avg(col):
        def f(r, L):
            off = _offset(f'{L[col]}{r}', n)
            return f'=IF(COUNT({off})<{n},"",AVERAGE({off}))'
        return f

    def plusdi(r, L):
        a = f'{L["ATR (helper)"]}{r}'; pa = f'{L["PlusDM Avg (helper)"]}{r}'
        # No .fillna(0): an undefined ATR leaves DI blank rather than zero.
        return f'=IF(OR({a}="",{a}=0,{pa}=""),"",{pa}/{a}*100)'

    def minusdi(r, L):
        a = f'{L["ATR (helper)"]}{r}'; ma = f'{L["MinusDM Avg (helper)"]}{r}'
        return f'=IF(OR({a}="",{a}=0,{ma}=""),"",{ma}/{a}*100)'

    def dx(r, L):
        pdi = f'{L["+DI"]}{r}'; mdi = f'{L["-DI"]}{r}'
        return (f'=IF(OR({pdi}="",{mdi}=""),"",'
                f'IF(({pdi}+{mdi})=0,"",ABS({pdi}-{mdi})/({pdi}+{mdi})*100))')

    def adx(r, L):
        off = _offset(f'{L["DX (helper)"]}{r}', n)
        return f'=IF(COUNT({off})<{n},"",AVERAGE({off}))'

    def buyc(r, L):
        if r == 2:
            return "=FALSE"
        return _cross_guarded(f'{L["ADX (calc)"]}{r-1}', f'{L["ADX (calc)"]}{r}',
                              strong, BUYDIR, warm)

    def sellc(r, L):
        if r == 2:
            return "=FALSE"
        return _cross_guarded(f'{L["ADX (calc)"]}{r-1}', f'{L["ADX (calc)"]}{r}',
                              weak, SELLDIR, warm)

    return [
        ("Delta (helper)", True, delta),
        ("PlusDM (helper)", True, plusdm),
        ("MinusDM (helper)", True, minusdm),
        ("TR (helper)", True, tr),
        ("ATR (helper)", True, _avg("TR (helper)")),
        ("PlusDM Avg (helper)", True, _avg("PlusDM (helper)")),
        ("MinusDM Avg (helper)", True, _avg("MinusDM (helper)")),
        ("+DI", False, plusdi),
        ("-DI", False, minusdi),
        ("DX (helper)", True, dx),
        ("ADX (calc)", False, adx),
        ("Buy Condition", False, buyc),
        ("Sell Condition", False, sellc),
    ]


def _heikin_ashi_spec(S):
    warm = "1"                      # row 0's HA Open is a seed, not a candle
    trans = S.get("transition_only", "TRUE")

    def haclose(r, L):
        return f'={L["price"]}{r}'

    def haopen(r, L):
        if r == 2:
            return f'={L["price"]}{r}'
        prevopen = f'{L["HA Open"]}{r-1}'; prevclose = f'{L["HA Close (calc)"]}{r-1}'
        return f'=({prevopen}+{prevclose})/2'

    def buyt(r, L):
        o = f'{L["HA Open"]}{r}'
        return f'={o}*(1+IF({BUYDIR}="above",1,-1)*{BUYPCT}/100)'

    def sellt(r, L):
        o = f'{L["HA Open"]}{r}'
        return f'={o}*(1+IF({SELLDIR}="above",1,-1)*{SELLPCT}/100)'

    def buyc(r, L):
        c = f'{L["HA Close (calc)"]}{r}'; bt = f'{L["Buy Threshold"]}{r}'
        now = f'IF({BUYDIR}="above",{c}>{bt},{c}<{bt})'
        if r == 2:
            return "=FALSE"
        pc = f'{L["HA Close (calc)"]}{r-1}'; pbt = f'{L["Buy Threshold"]}{r-1}'
        prev = f'IF({BUYDIR}="above",{pc}>{pbt},{pc}<{pbt})'
        return _gate(warm, f'IF({trans},AND({now},NOT({prev})),{now})')

    def sellc(r, L):
        c = f'{L["HA Close (calc)"]}{r}'; st = f'{L["Sell Threshold"]}{r}'
        now = f'IF({SELLDIR}="below",{c}<{st},{c}>{st})'
        if r == 2:
            return "=FALSE"
        pc = f'{L["HA Close (calc)"]}{r-1}'; pst = f'{L["Sell Threshold"]}{r-1}'
        prev = f'IF({SELLDIR}="below",{pc}<{pst},{pc}>{pst})'
        return _gate(warm, f'IF({trans},AND({now},NOT({prev})),{now})')

    return [
        ("HA Close (calc)", False, haclose),
        ("HA Open", False, haopen),
        ("Buy Threshold", False, buyt),
        ("Sell Threshold", False, sellt),
        ("Buy Condition", False, buyc),
        ("Sell Condition", False, sellc),
    ]


INDICATOR_SPECS = {
    "Simple Moving Average":      _sma_spec,
    "Exponential Moving Average": _ema_spec,
    "Stochastic Oscillator":      _stochastic_spec,
    "MACD":                       _macd_spec,
    "Bollinger Bands":            _bollinger_spec,
    "Relative Strength Index":    _rsi_spec,
    "Fibonacci Retracement":      _fibonacci_spec,
    "Standard Deviation":         _stddev_spec,
    "ADX":                        _adx_spec,
    "Heikin Ashi":                _heikin_ashi_spec,
}


def build_workbook(
    df_raw,
    indicator_name: str,
    price_col: str,
    window: int,
    buy_pct: float,
    sell_pct: float,
    buy_direction: str,
    sell_direction: str,
    repeat_flag: bool = False,
    params: dict | None = None,
):
    """Returns an openpyxl Workbook with every calculated cell as a live formula.

    `params` carries the indicator settings the sidebar collected. They are
    written to the Settings sheet and referenced by cell, so the workbook and
    the app agree, and the values stay editable in Excel.
    """
    if indicator_name not in INDICATOR_SPECS:
        raise ValueError(f"Unknown indicator: {indicator_name}")

    params = dict(params or {})
    wb = Workbook()
    S = _build_settings_sheet(
        wb, window, buy_pct, sell_pct, buy_direction, sell_direction,
        repeat_flag, params, indicator_name,
    )
    ws = wb.create_sheet("Results")

    raw_cols = [c for c in RAW_COL_ORDER if c in df_raw.columns]
    spec = INDICATOR_SPECS[indicator_name](S)
    calc_names = [name for name, _, _ in spec] + ["Position (calc)", "Action (calc)", "Status"]
    all_cols = raw_cols + calc_names

    L = {name: get_column_letter(i) for i, name in enumerate(all_cols, start=1)}
    L["price"] = L[price_col]

    n = len(df_raw)
    last_row = n + 1

    # ── header row ──────────────────────────────────────────────────────────
    hidden_set = {name for name, hidden, _ in spec if hidden}
    for col_idx, name in enumerate(all_cols, start=1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        if name in hidden_set:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    # ── raw data rows (static values, blue font = inputs) ──────────────────
    has_action = "Action" in df_raw.columns
    for i, (_, row) in enumerate(df_raw.iterrows()):
        r = i + 2
        for name in raw_cols:
            val = row.get(name, None)
            c = ws.cell(row=r, column=all_cols.index(name) + 1)
            if name == "Transaction Time" and pd_isna_safe(val) is False:
                try:
                    c.value = val.to_pydatetime() if hasattr(val, "to_pydatetime") else val
                    c.number_format = "yyyy-mm-dd hh:mm:ss"
                except Exception:
                    c.value = str(val)
            else:
                c.value = None if pd_isna_safe(val) else val
            c.font = BLUE

    # ── calculated columns: write formulas row by row ──────────────────────
    for r in range(2, last_row + 1):
        for name, hidden, fn in spec:
            c = ws.cell(row=r, column=all_cols.index(name) + 1, value=fn(r, L))
            c.font = GREY if hidden else BLACK

        prev_pos_ref = '"Out"' if r == 2 else f'{L["Position (calc)"]}{r-1}'
        buy_cell = f'{L["Buy Condition"]}{r}'
        sell_cell = f'{L["Sell Condition"]}{r}'
        pos_f, act_f = _state_machine_formulas(r, buy_cell, sell_cell, prev_pos_ref, repeat_flag)
        ws.cell(row=r, column=all_cols.index("Position (calc)") + 1, value=pos_f).font = BLACK
        ws.cell(row=r, column=all_cols.index("Action (calc)") + 1, value=act_f).font = BLACK

        status_f = (_status_formula(f'{L["Action"]}{r}', f'{L["Action (calc)"]}{r}')
                    if has_action else "N/A")
        ws.cell(row=r, column=all_cols.index("Status") + 1, value=status_f).font = BLACK

    # ── cosmetics ────────────────────────────────────────────────────────────
    for col_idx, name in enumerate(all_cols, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(12, min(22, len(name) + 4))
    ws.freeze_panes = "A2"

    # ── summary block (formula-driven) ─────────────────────────────────────
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Metric"; summary["B1"] = "Value"
    for cell in ("A1", "B1"):
        summary[cell].font = HDR_FONT
        summary[cell].fill = HDR_FILL
    action_rng = f'Results!{L["Action (calc)"]}2:{L["Action (calc)"]}{last_row}'
    status_rng = f'Results!{L["Status"]}2:{L["Status"]}{last_row}'
    rows = [
        ("Total Rows", f"=COUNTA({action_rng})"),
        ("Buy Signals", f'=COUNTIF({action_rng},"Buy")'),
        ("Sell Signals", f'=COUNTIF({action_rng},"Sell")'),
        ("Pass", f'=COUNTIF({status_rng},"Pass")'),
        ("Fail", f'=COUNTIF({status_rng},"Fail")'),
        ("N/A", f'=COUNTIF({status_rng},"N/A")'),
    ]
    for i, (label, formula) in enumerate(rows, start=2):
        summary.cell(row=i, column=1, value=label).font = Font(bold=True)
        summary.cell(row=i, column=2, value=formula).font = BLACK
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 14

    return wb


def pd_isna_safe(val):
    try:
        import pandas as pd
        return bool(pd.isna(val))
    except Exception:
        return val is None